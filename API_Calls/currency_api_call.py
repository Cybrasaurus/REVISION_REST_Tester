import requests
import openpyxl 
import os
import json
import datetime
from API_Calls.Currency_api_data_to_df import saver, opener, to_df, get_df_data


def get_currency_info(country):
    """
    This function calls the CurrencyScoop API.
    The returned object is converted to a json, in the metric System
    :param country: the name of the country 
    :return: a touple consisting of the 
        result1 : dict of basic infos
        result2 : latest currency rate (float or int, rounded)
        result3 : average cur rate, specify int for days (float or int, rounded)
        percent_string (result2, result3) : string, mit vorzeichen und % 
    """

    # 
    url1 = "https://currencyscoop.p.rapidapi.com/currencies"
    url2 = "https://currencyscoop.p.rapidapi.com/latest?base=EUR"      #u can specify base currency here
    headers = {
        "X-RapidAPI-Host": "currencyscoop.p.rapidapi.com",
	    "X-RapidAPI-Key": "14b2a42e1cmsh8b4668038092d99p17c629jsnbdd3c14469b1"}
    
    response1 = requests.request ("GET", url1, headers=headers)
    response2 = requests.request ("GET", url2, headers=headers)

    cc = get_curcode_2(country=country)
    if cc == None:
        return ["Error", "Could not Find Country in Database"]
    else:

        res1 = response1.json ()
        res2 = response2.json ()

        cod1 = res1 ["meta"]["code"]
        cod2 = res2 ["meta"]["code"]
        result1 = res1 ["response"]["fiats"][cc]
        result2 = round (res2 ["response"]["rates"][cc], 2)

        save_timeline (7)
        result3 = df_data (7, cc)       #specify days

        if type(cc) != str:
            return "Couldn't get currency code of " + country

        elif cod1 != 200 or cod2 != 200:
            return f"Error with the Currency API Call: HTTP response codes {cod1} and {cod2}, take a look on https://currencyscoop.com/api-documentation"

        elif type (result1) != dict:
            return f"Couldn't find any information on currency code {cc}"

        else:
            return result1, result2, result3, percent_string (result2, result3)


def get_curcode (country):
    #xlsx created with data from https://de.iban.com/currency-codes, changed it a bit myself
    currency_code_file = openpyxl.load_workbook(os.getcwd() + '/API_Calls/currency-codes.xlsx').get_sheet_by_name("Sheet1")
    try:
        for i in range(1,currency_code_file.max_row):
            if currency_code_file.cell(row=i, column=1).value == country.upper():
                return currency_code_file.cell(row=i, column=3).value        
    except:
        return None

def get_curcode_2(country):
    try:
        country_list = opener("API_Calls/Cleaned_IATA_V3_and_Curr")
        cur_code = country_list[country]["Currency_data"]["Currency Short"].upper()
        return cur_code
    except KeyError:
        return None


def save_timeline (days):
    url3 = "https://currencyscoop.p.rapidapi.com/historical?base=EUR"
    headers = {
        "X-RapidAPI-Host": "currencyscoop.p.rapidapi.com",
	    "X-RapidAPI-Key": "14b2a42e1cmsh8b4668038092d99p17c629jsnbdd3c14469b1"}
    for i in range (1, days+1):
        Previous_Date = datetime.datetime.today() - datetime.timedelta(days=i)
        Previous_Date_Formatted = Previous_Date.strftime ("%Y-%m-%d") # format the date to ddmmyyyy
        #print (str(Previous_Date_Formatted))          
        querydate = {"date":str(Previous_Date_Formatted)}
        response3 = requests.request("GET", url3, headers=headers, params=querydate)
        res3 = response3.json()
        saver(f"data{i}", res3)

def df_data (amount_ds, cc):
    df1 = to_df(amount_ds)
    average_cur_rate = get_df_data (df1, cc)
    return round (average_cur_rate, 2)

def percent_string (result2, result3):
    percentage_increase = round ((result2 - result3) / result3 * 100, 2)
    if percentage_increase > 0:
        vz = '+'
    else: 
        vz = ''
    s = vz + str(percentage_increase) + ' %'
    return s

if __name__ == "__main__":
    print(get_curcode_2("United States"))